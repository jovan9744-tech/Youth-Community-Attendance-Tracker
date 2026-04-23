from flask import Flask, request, jsonify, send_file, render_template, session, redirect, url_for
from flask_cors import CORS
import sqlite3
import os
from datetime import datetime, date
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import hashlib
import secrets

app = Flask(__name__, static_folder="static")
app.secret_key = "youth-tracker-secret-2024-xk9m"
CORS(app, supports_credentials=True)

DB_PATH = "database.db"

def hash_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL UNIQUE,
            count INTEGER NOT NULL,
            type TEXT NOT NULL CHECK(type IN ('Normal', 'Event', 'Group Activity')),
            sub_type TEXT,
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin', 'member')),
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)
    existing = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    if existing == 0:
        conn.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                     ("admin", hash_password("youth2024"), "admin"))
        conn.execute("INSERT INTO users (username, password, role) VALUES (?, ?, ?)",
                     ("member", hash_password("youth123"), "member"))
    conn.commit()
    conn.close()

#auth decorators
def require_login(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user"):
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated

def require_admin(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        user = session.get("user")
        if not user:
            if request.is_json or request.path.startswith("/api/"):
                return jsonify({"error": "Unauthorized"}), 401
            return redirect(url_for("login_page"))
        if user.get("role") != "admin":
            return jsonify({"error": "Akses ditolak. Hanya Admin yang bisa melakukan ini."}), 403
        return f(*args, **kwargs)
    return decorated

#page routes
@app.route("/login")
def login_page():
    if session.get("user"):
        return redirect(url_for("index"))
    return render_template("login.html")

@app.route("/")
@require_login
def index():
    return render_template("index.html", user=session.get("user"))

#auth api
@app.route("/api/auth/login", methods=["POST"])
def login():
    data = request.json
    username = data.get("username", "").strip()
    password = data.get("password", "")

    if not username or not password:
        return jsonify({"error": "Username dan password wajib diisi"}), 400

    conn = get_db()
    user = conn.execute(
        "SELECT * FROM users WHERE username=? AND password=?",
        (username, hash_password(password))
    ).fetchone()
    conn.close()

    if not user:
        return jsonify({"error": "Username atau password salah"}), 401

    session.permanent = True
    session["user"] = {"id": user["id"], "username": user["username"], "role": user["role"]}
    return jsonify({"username": user["username"], "role": user["role"]})

@app.route("/api/auth/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"message": "Logged out"})

@app.route("/api/auth/me", methods=["GET"])
def me():
    user = session.get("user")
    if not user:
        return jsonify({"error": "Not logged in"}), 401
    return jsonify(user)

@app.route("/api/auth/change-password", methods=["POST"])
@require_admin
def change_password():
    data = request.json
    target = data.get("username", "admin")
    new_pw = data.get("new_password", "")
    if len(new_pw) < 6:
        return jsonify({"error": "Password minimal 6 karakter"}), 400
    conn = get_db()
    conn.execute("UPDATE users SET password=? WHERE username=?", (hash_password(new_pw), target))
    conn.commit()
    conn.close()
    return jsonify({"message": f"Password {target} berhasil diubah"})


#for attendance api
@app.route("/api/attendance", methods=["GET"])
@require_login
def get_attendance():
    conn = get_db()
    rows = conn.execute("SELECT * FROM attendance ORDER BY date DESC").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/api/attendance", methods=["POST"])
@require_admin
def add_attendance():
    data = request.json
    for f in ["date", "count", "type"]:
        if f not in data:
            return jsonify({"error": f"Field '{f}' is required"}), 400
    try:
        d = datetime.strptime(data["date"], "%Y-%m-%d")
        if d.weekday() != 5:
            return jsonify({"error": "Tanggal harus hari Sabtu"}), 400
    except ValueError:
        return jsonify({"error": "Format tanggal tidak valid"}), 400

    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO attendance (date, count, type, sub_type, notes) VALUES (?, ?, ?, ?, ?)",
            (data["date"], data["count"], data["type"],
             data.get("sub_type", ""), data.get("notes", ""))
        )
        conn.commit()
        row = conn.execute("SELECT * FROM attendance WHERE date = ?", (data["date"],)).fetchone()
        conn.close()
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"error": "Data untuk tanggal ini sudah ada"}), 409

@app.route("/api/attendance/<int:id>", methods=["PUT"])
@require_admin
def update_attendance(id):
    data = request.json
    conn = get_db()
    row = conn.execute("SELECT * FROM attendance WHERE id = ?", (id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Data tidak ditemukan"}), 404
    conn.execute(
        "UPDATE attendance SET count=?, type=?, sub_type=?, notes=? WHERE id=?",
        (data.get("count", row["count"]), data.get("type", row["type"]),
         data.get("sub_type", row["sub_type"]), data.get("notes", row["notes"]), id)
    )
    conn.commit()
    updated = conn.execute("SELECT * FROM attendance WHERE id = ?", (id,)).fetchone()
    conn.close()
    return jsonify(dict(updated))

@app.route("/api/attendance/<int:id>", methods=["DELETE"])
@require_admin
def delete_attendance(id):
    conn = get_db()
    row = conn.execute("SELECT * FROM attendance WHERE id = ?", (id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Data tidak ditemukan"}), 404
    conn.execute("DELETE FROM attendance WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    return jsonify({"message": "Data berhasil dihapus"})

@app.route("/api/stats", methods=["GET"])
@require_login
def get_stats():
    conn = get_db()
    rows = conn.execute("SELECT * FROM attendance ORDER BY date ASC").fetchall()
    data = [dict(r) for r in rows]
    conn.close()

    if not data:
        return jsonify({
            "monthly": [], "yearly": [], "by_type": [],
            "summary": {"total_services": 0, "avg_monthly": 0, "highest": 0, "growth_pct": None}
        })

    df = pd.DataFrame(data)
    df["date"] = pd.to_datetime(df["date"])
    df["year"] = df["date"].dt.year
    df["month"] = df["date"].dt.month
    df["month_label"] = df["date"].dt.strftime("%b %Y")
    df["year_label"] = df["year"].astype(str)

    monthly = df.groupby(["year", "month", "month_label"])["count"].agg(
        total="sum", avg="mean", services="count"
    ).reset_index().sort_values(["year", "month"])

    yearly = df.groupby(["year", "year_label"])["count"].agg(
        total="sum", avg="mean", services="count"
    ).reset_index().sort_values("year")

    by_type = df.groupby("type")["count"].agg(total="sum", services="count").reset_index()

    now = datetime.now()
    this_month = df[(df["year"] == now.year) & (df["month"] == now.month)]
    lm = df[(df["year"] == now.year - 1) & (df["month"] == 12)] if now.month == 1 \
        else df[(df["year"] == now.year) & (df["month"] == now.month - 1)]

    tm_total = int(this_month["count"].sum()) if not this_month.empty else 0
    lm_total = int(lm["count"].sum()) if not lm.empty else 0
    growth_pct = round(((tm_total - lm_total) / lm_total) * 100, 1) if lm_total > 0 else None

    return jsonify({
        "monthly": monthly.to_dict("records"),
        "yearly": yearly.to_dict("records"),
        "by_type": by_type.to_dict("records"),
        "summary": {
            "total_services": len(df),
            "avg_monthly": round(float(df.groupby(["year","month"])["count"].sum().mean()), 1),
            "highest": int(df["count"].max()),
            "highest_date": df.loc[df["count"].idxmax(), "date"].strftime("%d %b %Y"),
            "this_month_total": tm_total,
            "last_month_total": lm_total,
            "growth_pct": growth_pct
        }
    })

@app.route("/api/export", methods=["GET"])
@require_login
def export_excel():
    conn = get_db()
    rows = conn.execute("SELECT * FROM attendance ORDER BY date ASC").fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Kehadiran"
    header_fill = PatternFill("solid", fgColor="1a1a1a")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(left=Side(style="thin", color="DDDDDD"), right=Side(style="thin", color="DDDDDD"),
                  top=Side(style="thin", color="DDDDDD"), bottom=Side(style="thin", color="DDDDDD"))

    headers = ["No", "Tanggal", "Hari", "Tipe Ibadah", "Sub-Tipe", "Jumlah Hadir", "Catatan"]
    col_widths = [5, 15, 12, 18, 20, 15, 30]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center; cell.border = thin
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30

    type_colors = {"Normal": "F0F0F0", "Event": "FFF3CD", "Group Activity": "D4EDDA"}
    for idx, row in enumerate(rows, 2):
        d = datetime.strptime(row["date"], "%Y-%m-%d")
        day_id = ["Senin","Selasa","Rabu","Kamis","Jumat","Sabtu","Minggu"][d.weekday()]
        row_fill = PatternFill("solid", fgColor=type_colors.get(row["type"], "FFFFFF"))
        for col, val in enumerate([idx-1, row["date"], day_id, row["type"],
                                    row["sub_type"] or "-", row["count"], row["notes"] or "-"], 1):
            cell = ws.cell(row=idx, column=col, value=val)
            cell.fill = row_fill; cell.border = thin
            cell.alignment = center if col != 7 else Alignment(horizontal="left", vertical="center")

    output = BytesIO()
    wb.save(output); output.seek(0)
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"rekap_youth_{datetime.now().strftime('%Y%m%d')}.xlsx")

@app.route("/api/import", methods=["POST"])
@require_admin
def import_excel():
    if "file" not in request.files:
        return jsonify({"error": "Tidak ada file yang diunggah"}), 400
    file = request.files["file"]
    if not file.filename.endswith((".xlsx", ".xls")):
        return jsonify({"error": "File harus berformat Excel"}), 400
    try:
        df = pd.read_excel(file)
        df.columns = [str(c).strip().lower() for c in df.columns]
        col_map = {}
        for col in df.columns:
            if any(k in col for k in ["tanggal", "date", "tgl"]): col_map["date"] = col
            elif any(k in col for k in ["jumlah", "hadir", "count", "total"]): col_map["count"] = col
            elif any(k in col for k in ["tipe", "type", "jenis"]): col_map["type"] = col
            elif any(k in col for k in ["sub", "kategori"]): col_map["sub_type"] = col
            elif any(k in col for k in ["catatan", "notes", "keterangan"]): col_map["notes"] = col

        if "date" not in col_map or "count" not in col_map:
            return jsonify({"error": "Kolom Tanggal dan Jumlah Hadir tidak ditemukan"}), 400

        preview, errors = [], []
        for idx, row in df.iterrows():
            try:
                raw_date = row[col_map["date"]]
                if pd.isna(raw_date): continue
                d = pd.to_datetime(raw_date, dayfirst=True) if isinstance(raw_date, str) else pd.to_datetime(raw_date)
                itype = str(row.get(col_map.get("type",""), "Normal")).strip()
                if itype not in ["Normal", "Event", "Group Activity"]: itype = "Normal"
                sub = str(row.get(col_map.get("sub_type",""), "")).strip() if col_map.get("sub_type") else ""
                notes = str(row.get(col_map.get("notes",""), "")).strip() if col_map.get("notes") else ""
                for v in ["nan", "NaN", "None"]:
                    if sub == v: sub = ""
                    if notes == v: notes = ""
                preview.append({"date": d.strftime("%Y-%m-%d"), "count": int(row[col_map["count"]]),
                                 "type": itype, "sub_type": sub, "notes": notes})
            except Exception as e:
                errors.append(f"Baris {idx+2}: {str(e)}")
        return jsonify({"preview": preview, "errors": errors, "total": len(preview)})
    except Exception as e:
        return jsonify({"error": f"Gagal membaca file: {str(e)}"}), 500

@app.route("/api/import/confirm", methods=["POST"])
@require_admin
def confirm_import():
    data = request.json
    records = data.get("records", [])
    skip_duplicates = data.get("skip_duplicates", True)
    conn = get_db()
    inserted = skipped = 0
    for rec in records:
        try:
            conn.execute("INSERT INTO attendance (date, count, type, sub_type, notes) VALUES (?,?,?,?,?)",
                         (rec["date"], rec["count"], rec["type"], rec.get("sub_type",""), rec.get("notes","")))
            inserted += 1
        except sqlite3.IntegrityError:
            if skip_duplicates: skipped += 1
            else:
                conn.execute("UPDATE attendance SET count=?,type=?,sub_type=?,notes=? WHERE date=?",
                             (rec["count"], rec["type"], rec.get("sub_type",""), rec.get("notes",""), rec["date"]))
                inserted += 1
    conn.commit(); conn.close()
    return jsonify({"inserted": inserted, "skipped": skipped})

if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5000)
